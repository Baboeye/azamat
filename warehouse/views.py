from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Q, F
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.utils import timezone
from datetime import timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill

from .models import Material, Contractor, Section, AuditLog
from .forms import MaterialForm, ContractorForm, SectionForm


@login_required
def dashboard(request):
    """Улучшенная панель управления"""
    materials = Material.objects.select_related('section').all()
    total_materials = materials.count()
    total_quantity = sum(m.quantity for m in materials)
    low_stock = materials.filter(quantity__lte=F('min_stock')).count()
    
    # Данные для графика
    chart_data = {
        'labels': [m.name[:20] for m in materials[:10]],
        'quantities': [m.quantity for m in materials[:10]],
        'colors': ['#ff416c' if m.is_low_stock else '#00eaff' for m in materials[:10]]
    }
    
    # Последние операции
    recent_ops = AuditLog.objects.select_related('material', 'user').order_by('-created_at')[:10]
    
    # Критические остатки
    critical = materials.filter(quantity__lte=F('min_stock')).order_by('quantity')[:5]
    
    return render(request, 'warehouse/dashboard.html', {
        'total_materials': total_materials,
        'total_quantity': total_quantity,
        'low_stock': low_stock,
        'chart_data': chart_data,
        'recent_operations': recent_ops,
        'critical_materials': critical,
        'now': timezone.now()
    })


@login_required
def materials(request):
    """Список материалов с фильтрацией и поиском"""
    queryset = Material.objects.select_related('section', 'contractor').all()
    
    # Поиск
    search = request.GET.get('search')
    if search:
        queryset = queryset.filter(
            Q(name__icontains=search) | Q(article__icontains=search)
        )
    
    # Фильтр по секции
    section_id = request.GET.get('section')
    if section_id:
        queryset = queryset.filter(section_id=section_id)
    
    # Фильтр по статусу остатка
    stock_status = request.GET.get('stock_status')
    if stock_status == 'low':
        queryset = queryset.filter(quantity__lte=F('min_stock'))
    elif stock_status == 'normal':
        queryset = queryset.filter(quantity__gt=F('min_stock'))
    
    # Добавляем расчет процента для каждого материала
    materials_list = list(queryset)
    for m in materials_list:
        if m.min_stock > 0:
            m.stock_percent = min((m.quantity / (m.min_stock * 2)) * 100, 100)
        else:
            m.stock_percent = 100 if m.quantity > 0 else 0
    
    # Пагинация
    paginator = Paginator(materials_list, 12)
    page = request.GET.get('page')
    materials = paginator.get_page(page)
    
    return render(request, 'warehouse/materials.html', {
        'materials': materials,
        'total_count': queryset.count(),
        'sections': Section.objects.all()
    })


@login_required
def material_detail(request, pk):
    """Детальная информация о материале"""
    material = get_object_or_404(
        Material.objects.select_related('section', 'contractor'), 
        pk=pk
    )
    
    # История операций
    receipts = material.receipts.order_by('-created_at')[:10]
    issues = material.issues.order_by('-created_at')[:10]
    scraps = material.scraps.order_by('-created_at')[:10]
    
    # Статистика
    total_receipts = material.receipts.aggregate(total=Sum('quantity'))['total'] or 0
    total_issues = material.issues.aggregate(total=Sum('quantity'))['total'] or 0
    
    # Оборот за 30 дней
    from_date = timezone.now() - timedelta(days=30)
    turnover_receipts = material.receipts.filter(
        created_at__gte=from_date
    ).aggregate(total=Sum('quantity'))['total'] or 0
    turnover_issues = material.issues.filter(
        created_at__gte=from_date
    ).aggregate(total=Sum('quantity'))['total'] or 0
    
    # Добавляем свойство для процента остатка (для прогресс-бара)
    if material.min_stock > 0:
        stock_percent = min((material.quantity / (material.min_stock * 2)) * 100, 100)
    else:
        stock_percent = 100 if material.quantity > 0 else 0
    
    return render(request, 'warehouse/material_detail.html', {
        'material': material,
        'receipts': receipts,
        'issues': issues,
        'scraps': scraps,
        'total_receipts': total_receipts,
        'total_issues': total_issues,
        'turnover': {'receipts': turnover_receipts, 'issues': turnover_issues},
        'stock_percent': stock_percent,
    })


@login_required
def material_add(request):
    """Создание материала"""
    if request.method == 'POST':
        form = MaterialForm(request.POST, request.FILES)
        if form.is_valid():
            material = form.save()
            AuditLog.objects.create(
                action='create', 
                material=material, 
                user=request.user, 
                details='Создание материала'
            )
            messages.success(request, f'Материал "{material.name}" создан')
            return redirect('materials')
    else:
        form = MaterialForm()
    
    return render(request, 'warehouse/material_add.html', {
        'form': form,
        'sections': Section.objects.all(),
        'contractors': Contractor.objects.filter(contractor_type='supplier')
    })


@login_required
def material_edit(request, pk):
    """Редактирование материала"""
    material = get_object_or_404(Material, pk=pk)
    if request.method == 'POST':
        form = MaterialForm(request.POST, request.FILES, instance=material)
        if form.is_valid():
            form.save()
            AuditLog.objects.create(
                action='edit', 
                material=material, 
                user=request.user, 
                details='Редактирование материала'
            )
            messages.success(request, 'Материал обновлен')
            return redirect('material_detail', pk=pk)
    else:
        form = MaterialForm(instance=material)
    
    return render(request, 'warehouse/material_edit.html', {
        'form': form,
        'material': material,
        'sections': Section.objects.all(),
        'contractors': Contractor.objects.filter(contractor_type='supplier')
    })


@login_required
def material_delete(request, pk):
    """Удаление материала"""
    material = get_object_or_404(Material, pk=pk)
    if request.method == 'POST':
        AuditLog.objects.create(
            action='delete', 
            material=material, 
            user=request.user, 
            details=f'Удаление {material.name}'
        )
        material.delete()
        messages.success(request, 'Материал удален')
        return redirect('materials')
    return render(request, 'warehouse/material_confirm_delete.html', {'material': material})


@login_required
def material_receipt(request, pk):
    """Оформление прихода материала"""
    material = get_object_or_404(Material, pk=pk)
    if request.method == 'POST':
        quantity = float(request.POST.get('quantity', 0))
        if quantity > 0:
            material.quantity += quantity
            material.save()
            AuditLog.objects.create(
                action='receipt', 
                material=material, 
                quantity=quantity, 
                user=request.user
            )
            messages.success(request, f'Приход +{quantity} {material.unit} оформлен')
        return redirect('material_detail', pk=pk)
    return render(request, 'warehouse/receipt_form.html', {'material': material})


@login_required
def material_issue(request, pk):
    """Оформление расхода материала"""
    material = get_object_or_404(Material, pk=pk)
    if request.method == 'POST':
        quantity = float(request.POST.get('quantity', 0))
        if quantity > 0:
            if quantity > material.quantity:
                messages.error(
                    request, 
                    f'Недостаточно материала. Доступно: {material.quantity}'
                )
                return redirect('material_issue', pk=pk)
            material.quantity -= quantity
            material.save()
            AuditLog.objects.create(
                action='issue', 
                material=material, 
                quantity=-quantity, 
                user=request.user,
                details=f"Выдано: {request.POST.get('issued_to', '')}"
            )
            messages.success(request, f'Расход -{quantity} {material.unit} оформлен')
        return redirect('material_detail', pk=pk)
    return render(request, 'warehouse/issue_form.html', {'material': material})


@login_required
def material_scrap(request, pk):
    """Списание материала"""
    material = get_object_or_404(Material, pk=pk)
    if request.method == 'POST':
        quantity = float(request.POST.get('quantity', 0))
        reason = request.POST.get('reason', '')
        if quantity > 0:
            if quantity > material.quantity:
                messages.error(
                    request, 
                    f'Недостаточно материала. Доступно: {material.quantity}'
                )
                return redirect('material_scrap', pk=pk)
            material.quantity -= quantity
            material.save()
            AuditLog.objects.create(
                action='scrap', 
                material=material, 
                quantity=-quantity, 
                user=request.user,
                details=f'Причина: {reason}'
            )
            messages.warning(request, f'Списано {quantity} {material.unit}: {reason}')
        return redirect('material_detail', pk=pk)
    return render(request, 'warehouse/material_scrap.html', {'material': material})


@login_required
def contractors(request):
    """Список контрагентов"""
    suppliers = Contractor.objects.filter(contractor_type='supplier')
    customers = Contractor.objects.filter(contractor_type='customer')
    return render(request, 'warehouse/contractors.html', {
        'suppliers': suppliers, 
        'customers': customers
    })


@login_required
def add_contractor(request):
    """Добавление контрагента"""
    if request.method == 'POST':
        form = ContractorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Контрагент добавлен')
            return redirect('contractors')
    else:
        form = ContractorForm()
    return render(request, 'warehouse/add_contractor.html', {'form': form})


@login_required
def edit_contractor(request, pk):
    """Редактирование контрагента"""
    contractor = get_object_or_404(Contractor, pk=pk)
    if request.method == 'POST':
        form = ContractorForm(request.POST, instance=contractor)
        if form.is_valid():
            form.save()
            messages.success(request, 'Контрагент обновлен')
            return redirect('contractors')
    else:
        form = ContractorForm(instance=contractor)
    return render(request, 'warehouse/edit_contractor.html', {
        'form': form, 
        'contractor': contractor
    })


@login_required
def delete_contractor(request, pk):
    """Удаление контрагента"""
    contractor = get_object_or_404(Contractor, pk=pk)
    if request.method == 'POST':
        contractor.delete()
        messages.success(request, 'Контрагент удален')
        return redirect('contractors')
    return render(request, 'warehouse/contractor_confirm_delete.html', {'contractor': contractor})


@login_required
def warehouse(request):
    """Страница склада с секциями"""
    sections = Section.objects.prefetch_related('materials').all()
    return render(request, 'warehouse/warehouse.html', {'sections': sections})


@login_required
def add_section(request):
    """Добавление секции"""
    if request.method == 'POST':
        form = SectionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Секция добавлена')
            return redirect('warehouse')
    else:
        form = SectionForm()
    return render(request, 'warehouse/add_section.html', {'form': form})


@login_required
def delete_section(request, pk):
    """Удаление секции"""
    section = get_object_or_404(Section, pk=pk)
    if section.materials.exists():
        messages.error(request, 'Нельзя удалить секцию с материалами')
    else:
        section.delete()
        messages.success(request, 'Секция удалена')
    return redirect('warehouse')


@login_required
def reports(request):
    """Отчеты"""
    return render(request, 'warehouse/reports.html')


@login_required
def stock_report(request):
    """Отчет по остаткам"""
    materials = Material.objects.select_related('section').all()
    return render(request, 'warehouse/stock_report.html', {'materials': materials})


@login_required
def audit_log(request):
    """Журнал аудита"""
    logs = AuditLog.objects.select_related('material', 'user').order_by('-created_at')
    
    # Фильтр по действию
    action = request.GET.get('action')
    if action:
        logs = logs.filter(action=action)
    
    # Пагинация
    paginator = Paginator(logs, 20)
    page = request.GET.get('page')
    logs = paginator.get_page(page)
    
    return render(request, 'warehouse/audit_log.html', {
        'logs': logs,
        'action_choices': AuditLog.ACTION_CHOICES,
    })

@login_required
def api_materials(request):
    """API для данных графиков"""
    materials = Material.objects.all()[:20]
    return JsonResponse({
        'names': [m.name for m in materials],
        'quantities': [m.quantity for m in materials],
        'min_stocks': [m.min_stock for m in materials]
    })


@login_required
def export_materials_excel(request):
    """Экспорт материалов в Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Материалы"
    
    # Заголовки
    headers = ['Артикул', 'Название', 'Количество', 'Ед.изм', 'Секция', 'Мин.остаток', 'Статус']
    ws.append(headers)
    
    # Стили заголовков
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0077ff", end_color="0077ff", fill_type="solid")
    
    # Данные
    for material in Material.objects.select_related('section').all():
        status = 'Критично' if material.is_low_stock else 'Норма'
        ws.append([
            material.article,
            material.name,
            material.quantity,
            material.unit,
            material.section.name if material.section else '-',
            material.min_stock,
            status
        ])
    
    # Автоширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="materials.xlsx"'
    wb.save(response)
    return response